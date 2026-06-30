#!/usr/bin/env python3
"""notion_to_excel.py — Compila la parrilla de contenido exportada de Notion a un .xlsx.

Lee el export .zip MÁS RECIENTE de la carpeta notion/ (Notion exporta un zip
dentro de otro zip; el CSV principal se detecta por columnas, no por nombre).
Filtra por mes (ej. "Julio") y genera un Excel con:
  - Hoja TODO  (todas las publicaciones del mes, todos los tenants)
  - Una hoja por tenant/marca (según la columna Tenant del export)
  - Hoja _SIN-TENANT si hay publicaciones sin tenant asignado (aviso de QA)

Autocontenido: no depende de otros proyectos. Skill: /notion-to-excel

Uso:
  python3 notion_to_excel.py --mes Julio
  python3 notion_to_excel.py --mes Jul --zip "notion/notion_V4 (julio).zip"
  python3 notion_to_excel.py            # sin --mes: incluye TODOS los meses
"""
import os, sys, csv, io, zipfile, argparse, re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJ = os.path.dirname(os.path.abspath(__file__))
NOTION_DIR = os.path.join(PROJ, "notion")
OUT_DIR = os.path.join(PROJ, "output")

# Orden lógico de columnas (las del export de Notion). Si aparecen columnas nuevas
# en el export se añaden al final automáticamente, así nunca se pierde información.
COLS_BASE = ["Tenant", "Mes", "Pilar", "Tipo", "Estado", "Name", "Date",
             "Dirección de Arte", "Copy (ES)", "Copy (POR)", "Hashtags"]
WIDTHS = {"Tenant": 16, "Mes": 9, "Pilar": 14, "Tipo": 14, "Estado": 13,
          "Name": 24, "Date": 16, "Dirección de Arte": 55,
          "Copy (ES)": 55, "Copy (POR)": 55, "Hashtags": 38}

# Nombres amigables para las pestañas por tenant (opcional). Los tenants que no estén
# aquí usan su clave tal cual. Ejemplo: {"acme": "ACME (US)", "acme-mx": "ACME (México)"}
TENANT_NICE = {}

# Meses en español -> clave de 3 letras como las usa Notion ("Jul 26", "Jun 26")
MES_KEY = {"enero": "ene", "febrero": "feb", "marzo": "mar", "abril": "abr",
           "mayo": "may", "junio": "jun", "julio": "jul", "agosto": "ago",
           "septiembre": "sep", "setiembre": "sep", "octubre": "oct",
           "noviembre": "nov", "diciembre": "dic"}

HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPLEFT = Alignment(vertical="top", wrap_text=True)
ZEBRA = PatternFill("solid", fgColor="F3F4F6")


# ---------- carga del export de Notion ----------
def zip_vigente(notion_dir):
    """El .zip más reciente en notion/. Al subir un export nuevo se toma solo."""
    if not os.path.isdir(notion_dir):
        sys.exit(f"❌ No existe la carpeta de export: {notion_dir}")
    zips = [os.path.join(notion_dir, f) for f in os.listdir(notion_dir)
            if f.lower().endswith(".zip")]
    if not zips:
        sys.exit(f"❌ No hay ningún export .zip en {notion_dir}")
    return max(zips, key=os.path.getmtime)


def _read_csv_bytes(data):
    return list(csv.DictReader(io.StringIO(data.decode("utf-8-sig"))))


def load_export(zip_path):
    """Devuelve los registros del CSV principal (el que trae 'Dirección de Arte').
    Maneja el zip-dentro-de-zip típico de Notion y detecta el CSV por columnas."""
    candidates = []
    with zipfile.ZipFile(zip_path) as outer:
        inner_names = [n for n in outer.namelist() if n.lower().endswith(".zip")]
        if inner_names:
            with zipfile.ZipFile(io.BytesIO(outer.read(inner_names[0]))) as inner:
                for n in inner.namelist():
                    if n.lower().endswith(".csv"):
                        candidates.append(_read_csv_bytes(inner.read(n)))
        else:
            for n in outer.namelist():
                if n.lower().endswith(".csv"):
                    candidates.append(_read_csv_bytes(outer.read(n)))
    for rows in candidates:
        if rows and "Dirección de Arte" in rows[0]:
            return rows
    sys.exit("❌ No se encontró el CSV principal de posts en el export de Notion.")


def mes_match(valor_mes, filtro):
    """'Jul 26' coincide con filtro 'Julio'/'Jul'. Compara por clave de 3 letras."""
    if not filtro:
        return True
    f = filtro.strip().lower()
    key = MES_KEY.get(f, f[:3])
    return (valor_mes or "").strip().lower().startswith(key)


# ---------- construcción del Excel ----------
def col_order(rows):
    """Orden base + cualquier columna extra del export al final."""
    extra = [c for c in (rows[0].keys() if rows else []) if c not in COLS_BASE]
    return [c for c in COLS_BASE if rows and c in rows[0]] + extra


def fill_sheet(ws, data, cols):
    ws.append(cols)
    for ci, _ in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci)
        c.fill = HEAD_FILL; c.font = HEAD_FONT
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        c.border = BORDER
    for ri, r in enumerate(data, 2):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=ri, column=ci, value=(r.get(col) or "").strip())
            c.alignment = TOPLEFT; c.border = BORDER
            if ri % 2 == 0:
                c.fill = ZEBRA
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    from openpyxl.utils import get_column_letter
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(col, 18)


def sheet_name(raw):
    """Nombre de pestaña válido para Excel (<=31 chars, sin caracteres prohibidos)."""
    name = re.sub(r'[\[\]:*?/\\]', '-', raw)
    return name[:31]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--mes", default=None,
                    help="Mes a compilar (ej. Julio). Si se omite: todos los meses.")
    ap.add_argument("--zip", default=None, help="Ruta a un .zip concreto (por defecto: el más reciente).")
    ap.add_argument("--out", default=None, help="Ruta de salida .xlsx (por defecto: output/ con consecutivo).")
    ap.add_argument("--list-meses", action="store_true",
                    help="Solo lista los meses detectados en el export (para el menú interactivo).")
    a = ap.parse_args()

    zip_path = a.zip if a.zip else zip_vigente(NOTION_DIR)
    if not os.path.isabs(zip_path):
        zip_path = os.path.join(PROJ, zip_path)
    rows = load_export(zip_path)

    # Modo listado: imprime "Mes<TAB>cantidad" por línea, en orden cronológico.
    if a.list_meses:
        cnt = Counter((r.get("Mes") or "(sin mes)").strip() for r in rows)
        orden = list(MES_KEY.values())
        keyf = lambda m: (orden.index(m[:3].lower()) if m[:3].lower() in orden else 99, m)
        print(f"# Fuente: {os.path.basename(zip_path)}")
        for mes in sorted(cnt, key=keyf):
            print(f"{mes}\t{cnt[mes]}")
        return

    sel = [r for r in rows if mes_match(r.get("Mes"), a.mes)]
    if not sel:
        sys.exit(f"❌ No hay publicaciones para el mes '{a.mes}' en {os.path.basename(zip_path)}.")

    cols = col_order(sel)
    etiqueta = (a.mes or "TODOS").strip().capitalize()

    wb = Workbook()
    ws = wb.active
    ws.title = "TODO"
    fill_sheet(ws, sel, cols)

    # Hojas por tenant (en orden de aparición), más _SIN-TENANT si aplica
    tenants = []
    for r in sel:
        t = (r.get("Tenant") or "").strip()
        if t and t not in tenants:
            tenants.append(t)
    for t in tenants:
        sub = [r for r in sel if (r.get("Tenant") or "").strip() == t]
        fill_sheet(wb.create_sheet(title=sheet_name(TENANT_NICE.get(t, t))), sub, cols)

    sin_tenant = [r for r in sel if not (r.get("Tenant") or "").strip()]
    if sin_tenant:
        fill_sheet(wb.create_sheet(title="_SIN-TENANT"), sin_tenant, cols)

    # Salida con consecutivo (_01, _02, ...) para no pisar archivos previos
    out = a.out
    if not out:
        os.makedirs(OUT_DIR, exist_ok=True)
        base = f"parrilla_{etiqueta}"
        n = 1
        while os.path.exists(os.path.join(OUT_DIR, f"{base}_{n:02d}.xlsx")):
            n += 1
        out = os.path.join(OUT_DIR, f"{base}_{n:02d}.xlsx")
    else:
        os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb.save(out)

    # Resumen para el usuario
    print(f"✅ Excel generado: {out}")
    print(f"   Fuente: {os.path.basename(zip_path)}")
    print(f"   Mes: {etiqueta}  ·  Publicaciones: {len(sel)}")
    porten = dict(Counter((r.get('Tenant') or '(sin tenant)').strip() for r in sel))
    print(f"   Por tenant: {porten}")
    print(f"   Hojas: {wb.sheetnames}")
    if sin_tenant:
        print(f"   ⚠️  {len(sin_tenant)} publicación(es) SIN tenant -> hoja _SIN-TENANT (corregir en Notion).")


if __name__ == "__main__":
    main()
